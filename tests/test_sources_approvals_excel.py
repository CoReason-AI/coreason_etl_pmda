# Copyright (c) 2025 CoReason, Inc.
#
# This software is proprietary and dual-licensed.
# Licensed under the Prosperity Public License 3.0 (the "License").
# A copy of the license is available at https://prosperitylicense.com/versions/3.0.0
# For details, see the LICENSE file.
# Commercial use beyond a 30-day trial requires a separate license.
#
# Source Code: https://github.com/CoReason-AI/coreason_etl_pmda

import io
from unittest.mock import MagicMock, patch

import openpyxl
from coreason_etl_pmda.sources.approvals import approvals_source


def create_mock_excel(empty: bool = False, no_valid_rows: bool = False) -> bytes:
    """Creates a simple in-memory Excel file for testing."""
    wb = openpyxl.Workbook()
    ws = wb.active

    if empty:
        # Save empty
        pass
    elif no_valid_rows:
        ws.append(["RandomHeader", "Col2"])
        ws.append(["Val1", "Val2"])
    else:
        # Valid data
        # Headers
        ws.append(["販売名", "承認番号", "承認年月日", "一般的名称", "申請者氏名", "審査報告書"])
        # Row 1 (With Hyperlink)
        ws.cell(row=2, column=1, value="ExcelBrand")
        ws.cell(row=2, column=2, value="111111")
        ws.cell(row=2, column=3, value="R3.1.1")
        ws.cell(row=2, column=4, value="ExcelGen")
        ws.cell(row=2, column=5, value="ExcelCompany")

        cell_link = ws.cell(row=2, column=6, value="Report PDF")
        cell_link.hyperlink = "http://example.com/report.pdf"

        # Row 2 (Without Hyperlink, plain text URL)
        ws.append(["ExcelBrand2", "222222", "R3.1.2", "ExcelGen2", "ExcelCompany2", "http://example.com/report2.pdf"])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def test_approvals_source_excel_ingestion() -> None:
    """
    Test that approvals_source follows the 'Scrape List -> Download Excel' pattern.
    It should find the Excel link on the page, download it, and extract rows.
    """

    html_content = """
    <html>
        <body>
            <h1>List of Approvals</h1>
            <p>
                <a href="approvals_2024.xlsx">Download List (Excel)</a>
            </p>
        </body>
    </html>
    """

    excel_bytes = create_mock_excel()

    with patch("coreason_etl_pmda.sources.common.fetch_url") as mock_fetch:
        mock_page_resp = MagicMock()
        mock_page_resp.content = html_content.encode("utf-8")
        mock_page_resp.encoding = "utf-8"

        mock_excel_resp = MagicMock()
        mock_excel_resp.content = excel_bytes
        mock_excel_resp.encoding = None

        def side_effect(url: str, **kwargs: dict[str, object]) -> MagicMock:
            if url.endswith(".xlsx"):
                return mock_excel_resp
            return mock_page_resp

        mock_fetch.side_effect = side_effect

        # Run
        data = list(approvals_source(url="http://example.com/list"))

        # Assertions
        assert len(data) == 2

        # Check extraction logic
        row1 = next(r for r in data if r["source_id"] == "111111")
        assert row1["raw_payload"].get("review_report_url") == "http://example.com/report.pdf"


def test_approvals_source_excel_fallback_empty() -> None:
    """
    Test fallback to HTML if Excel is empty.
    """
    # Need at least 2 matching headers: 販売名, 承認番号
    html_content = """
    <html>
        <body>
            <a href="empty.xlsx">List</a>
            <table>
                <tr><th>販売名</th><th>承認番号</th></tr>
                <tr><td>Fallback_HTML_Data</td><td>999999</td></tr>
            </table>
        </body>
    </html>
    """

    excel_bytes = create_mock_excel(empty=True)

    with patch("coreason_etl_pmda.sources.common.fetch_url") as mock_fetch:
        mock_page_resp = MagicMock()
        mock_page_resp.content = html_content.encode("utf-8")
        mock_page_resp.encoding = "utf-8"

        mock_excel_resp = MagicMock()
        mock_excel_resp.content = excel_bytes

        def side_effect(url: str, **kwargs: dict[str, object]) -> MagicMock:
            if url.endswith(".xlsx"):
                return mock_excel_resp
            return mock_page_resp

        mock_fetch.side_effect = side_effect

        data = list(approvals_source())
        # Should fallback to HTML
        assert len(data) == 1
        assert data[0]["raw_payload"]["販売名"] == "Fallback_HTML_Data"


def test_approvals_source_excel_fallback_irrelevant() -> None:
    """
    Test fallback to HTML if Excel has no valid rows.
    """
    # Need at least 2 matching headers
    html_content = """
    <html>
        <body>
            <a href="irrelevant.xlsx">List</a>
            <table>
                <tr><th>販売名</th><th>承認番号</th></tr>
                <tr><td>Fallback_HTML_Data</td><td>888888</td></tr>
            </table>
        </body>
    </html>
    """

    excel_bytes = create_mock_excel(no_valid_rows=True)

    with patch("coreason_etl_pmda.sources.common.fetch_url") as mock_fetch:
        mock_page_resp = MagicMock()
        mock_page_resp.content = html_content.encode("utf-8")
        mock_page_resp.encoding = "utf-8"

        mock_excel_resp = MagicMock()
        mock_excel_resp.content = excel_bytes

        def side_effect(url: str, **kwargs: dict[str, object]) -> MagicMock:
            if url.endswith(".xlsx"):
                return mock_excel_resp
            return mock_page_resp

        mock_fetch.side_effect = side_effect

        data = list(approvals_source())
        assert len(data) == 1
        assert data[0]["raw_payload"]["販売名"] == "Fallback_HTML_Data"
